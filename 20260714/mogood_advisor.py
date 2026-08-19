# -*- coding: utf-8 -*-
"""
MoGood 買い物アドバイザー（ルールベース版・LLM実験用）

家庭食材リストとスーパー食材リストをExcelファイルから読み込み、
指定した料理を作るために「何をスーパーで買えばよいか」を、
個人の好みに応じてアドバイスする。

データファイル（このスクリプトと同じフォルダに置く）:
  - 家庭食材リスト.xlsx   （カテゴリ / 数量 / 単位 / 賞味期限）
  - スーパー食材リスト.xlsx（カテゴリ / 商品名 / 価格 / 内容量 / 単位 / パック数 / 産地 / 健康属性）
  - レシピリスト.xlsx     （料理名 / 材料名 / カテゴリ候補 / 必要量 / 単位）

必要ライブラリ:  pip3 install openpyxl

使い方:
  python3 mogood_advisor.py            → 肉じゃがを3つの好みで実行
  python3 mogood_advisor.py カレーライス → 指定料理を3つの好みで実行

好み（preference）:
  - price  : 低価格重視（共通単価に換算して最も割安な商品を選ぶ）
  - health : 健康重視（有機・減塩・無添加・国産を優先）
  - stock  : 家庭在庫の効率消費重視（代替食材で在庫を使い切り、買い物と余りを最小化）
"""

import os
import sys
import unicodedata
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

TODAY = date(2026, 7, 13)
EXPIRY_WARN_DAYS = 3  # この日数以内に期限が来る食材は「早めに使う」対象

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
HOME_XLSX = os.path.join(DATA_DIR, "家庭食材リスト.xlsx")
SUPER_XLSX = os.path.join(DATA_DIR, "スーパー食材リスト.xlsx")
RECIPE_XLSX = os.path.join(DATA_DIR, "レシピリスト.xlsx")


def _resolve(path):
    """日本語ファイル名のUnicode正規化差異(NFC/NFD)を吸収して実在パスを返す。
    OneDrive等の同期経由でファイル名がNFDになると、コード内のNFCリテラルと
    一致せずFileNotFoundになるため、ディレクトリ内をNFC比較で探し直す。"""
    if os.path.exists(path):
        return path
    d = os.path.dirname(path) or "."
    target = unicodedata.normalize("NFC", os.path.basename(path))
    try:
        for name in os.listdir(d):
            if unicodedata.normalize("NFC", name) == target:
                return os.path.join(d, name)
    except OSError:
        pass
    return path  # 見つからなければ元のパス（従来通りのエラーを出す）

# ----------------------------------------------------------------------
# 1. Excelファイルからのデータ読み込み
# ----------------------------------------------------------------------
def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v), "%Y/%m/%d").date()

def load_home_stock(path=HOME_XLSX):
    """家庭食材リスト.xlsx → [{category, amount, unit, expiry}, ...]"""
    ws = load_workbook(_resolve(path), data_only=True).active
    stock = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stock.append({
            "category": str(row[0]).strip(),
            "amount": float(row[1]),
            "unit": str(row[2]).strip(),
            "expiry": _to_date(row[3]),
        })
    return stock

def load_supermarket(path=SUPER_XLSX):
    """スーパー食材リスト.xlsx → [{category, name, price, size, unit, packs, origin, health_tags}, ...]"""
    ws = load_workbook(_resolve(path), data_only=True).active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tags = str(row[7]).strip() if row[7] else ""
        items.append({
            "category": str(row[0]).strip(),
            "name": str(row[1]).strip(),
            "price": float(row[2]),
            "size": float(row[3]),
            "unit": str(row[4]).strip(),
            "packs": int(row[5]),
            "origin": str(row[6]).strip(),
            "health_tags": [t.strip() for t in tags.replace("，", "、").split("、") if t.strip()],
        })
    return items

def load_recipes(path=RECIPE_XLSX):
    """レシピリスト.xlsx → {料理名: [{ingredient, categories, amount, unit}, ...]}
    カテゴリ候補は「、」区切り。先頭が本来の指定、2番目以降は代替候補。"""
    ws = load_workbook(_resolve(path), data_only=True).active
    recipes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        dish = str(row[0]).strip()
        cats = [c.strip() for c in str(row[2]).replace("，", "、").split("、") if c.strip()]
        recipes.setdefault(dish, []).append({
            "ingredient": str(row[1]).strip(),
            "categories": cats,
            "amount": float(row[3]),
            "unit": str(row[4]).strip(),
        })
    return recipes

HOME_STOCK = load_home_stock()
SUPERMARKET = load_supermarket()
RECIPES = load_recipes()

# ----------------------------------------------------------------------
# 3. 補助関数
# ----------------------------------------------------------------------
def usable_stock(category):
    """期限切れでない在庫量を返す"""
    return sum(s["amount"] for s in HOME_STOCK
               if s["category"] == category and s["expiry"] >= TODAY)

def expiring_soon():
    """期限が近い家庭在庫（フードロス警告用）"""
    limit = TODAY + timedelta(days=EXPIRY_WARN_DAYS)
    return [s for s in HOME_STOCK if TODAY <= s["expiry"] <= limit]

def unit_price(p):
    """共通単価: 1単位（g/ml/個/本/袋等）あたりの価格"""
    return p["price"] / (p["size"] * p["packs"])

def unit_price_label(p):
    if p["unit"] in ("g", "ml"):
        return f"100{p['unit']}あたり {unit_price(p)*100:.1f}円"
    return f"1{p['unit']}あたり {unit_price(p):.1f}円"

def health_score(p):
    score = 2 * len(p["health_tags"])
    o = p["origin"]
    if ("国産" in o) or o.endswith(("県産", "道産", "府産", "都産", "島産")) or o == "国内製造":
        score += 1
    return score

def candidates(categories):
    return [p for p in SUPERMARKET if p["category"] in categories]

# ----------------------------------------------------------------------
# 4. 好み別の商品選択ロジック
# ----------------------------------------------------------------------
def select_product(cands, need, preference):
    """
    不足量 need を満たす商品を好みに応じて1つ選び、(商品, 理由) を返す。
    容量が不足分に満たない商品は複数個購入を想定せず、1商品で足りるものを優先。
    """
    enough = [p for p in cands if p["size"] * p["packs"] >= need] or cands

    if preference == "price":
        # 共通単価が最安のもの。単価同率なら総額の安いもの
        best = min(enough, key=lambda p: (unit_price(p), p["price"]))
        reason = f"{unit_price_label(best)} と候補中最安"
    elif preference == "health":
        # 健康属性・国産を優先し、同点なら総額の安いもの（買いすぎ防止）
        best = max(enough, key=lambda p: (health_score(p), -p["price"]))
        tags = "・".join(best["health_tags"]) or "なし"
        reason = f"健康属性: {tags} / 産地: {best['origin']}"
    elif preference == "stock":
        # 余りが最小になる容量を選ぶ（フードロス最小化）。同点なら安いもの
        best = min(enough, key=lambda p: (p["size"] * p["packs"] - need, p["price"]))
        leftover = best["size"] * best["packs"] - need
        reason = f"必要量{need:g}{best['unit']}に対し余り{leftover:g}{best['unit']}で最小"
    else:
        raise ValueError(f"未知の好み: {preference}")
    return best, reason

# ----------------------------------------------------------------------
# 5. メイン: 購入アドバイス生成
# ----------------------------------------------------------------------
def advise(dish, preference):
    labels = {"price": "低価格重視", "health": "健康重視", "stock": "家庭在庫の効率消費重視"}
    lines = [f"\n{'='*62}", f"● 料理: {dish} ／ 好み: {labels[preference]}", f"{'='*62}"]
    total = 0

    for req in RECIPES[dish]:
        cats = req["categories"]
        need = req["amount"]

        # 在庫消費重視なら代替カテゴリの在庫も積極的に使う
        use_cats = cats if preference == "stock" else cats[:1]
        stock_used = []
        remaining = need
        for c in use_cats:
            s = usable_stock(c)
            if s > 0 and remaining > 0:
                use = min(s, remaining)
                stock_used.append((c, use, req["unit"]))
                remaining -= use

        if remaining <= 0:
            used = "、".join(f"{c}{a:g}{u}" for c, a, u in stock_used)
            lines.append(f"  ○ {req['ingredient']:　<6}: 購入不要（家庭在庫 {used} で充足）")
            continue

        cands = candidates(cats)
        if not cands:
            lines.append(f"  × {req['ingredient']}: スーパーに該当商品なし")
            continue

        best, reason = select_product(cands, remaining, preference)
        total += best["price"]
        stock_note = ""
        if stock_used:
            used = "、".join(f"{c}{a:g}{u}" for c, a, u in stock_used)
            stock_note = f"（在庫 {used} を使用し不足 {remaining:g}{req['unit']} 分を購入）"
        lines.append(f"  ★ {req['ingredient']:　<6}: 「{best['name']}」 {best['price']:g}円 {stock_note}")
        lines.append(f"      └ 理由: {reason}")

    lines.append(f"\n  ▶ 購入合計: {total:g}円")

    exp = expiring_soon()
    if exp:
        items = "、".join(f"{s['category']}（{s['expiry'].strftime('%m/%d')}期限）" for s in exp)
        lines.append(f"  ⚠ 期限間近の在庫: {items} → 早めの消費を推奨")
        if preference == "stock":
            lines.append("     今回のレシピでこれらを優先的に使う選択をしています。")
    return "\n".join(lines)

# ----------------------------------------------------------------------
if __name__ == "__main__":
    dish = sys.argv[1] if len(sys.argv) > 1 else "肉じゃが"
    if dish not in RECIPES:
        raise SystemExit(f"レシピ未登録: {dish}（登録済み: {'、'.join(RECIPES)}）")
    print("MoGood 買い物アドバイザー（実験）  基準日:", TODAY)
    print(f"家庭在庫 {len(HOME_STOCK)}品 / スーパー商品 {len(SUPERMARKET)}品 をExcelから読み込みました")
    for pref in ["price", "health", "stock"]:
        print(advise(dish, pref))
